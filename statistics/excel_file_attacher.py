
from .file_sys_helper import FileSysHelper
from openpyxl import Workbook, load_workbook


class ExcelFileAttacher:
    """
    ExcelFileAttacher class is used to handle Excel file operations.
    """

    def __init__(self, file_path):
        self.file_path = file_path
        self.file_sys_helper = FileSysHelper(file_path)

    def attach_new_row(self, trade_dict: dict[str, str | float]):
        row = list(trade_dict.values())
        file_exists = self.file_sys_helper.file_exists()

        if not file_exists:
            workbook = Workbook()
            sheet = workbook.active
            # Headers
            headers = trade_dict.keys()
            sheet.append(list(headers))
            # Row
            sheet.append(row)
            workbook.save(self.file_path)

        if file_exists:
            workbook = load_workbook(self.file_path)
            sheet = workbook.active
            # Row
            sheet.append(row)
            workbook.save(self.file_path)
